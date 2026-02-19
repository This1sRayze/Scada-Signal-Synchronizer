import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import re
import subprocess
import os
import sys

class SCADASyncUI:
    def __init__(self, root):
        """Initialize the SCADA Signal Synchronizer UI application"""
        self.root = root
        self.root.title("SCADA Signal Synchronizer")
        self.root.geometry("500x600")
        self.root.configure(bg='#f0f0f0')
        
        # Instance variables for file and output management
        self.excel_file = None  # Path to selected Excel file
        self.area_descriptions = {}  # Dictionary to store descriptions from area sheets
        self.output_path = None  # User-selected output directory path
        self.output_name = None  # User-provided output filename
        
        # Setup UI components
        self.setup_styles()
        self.setup_ui()
    
    def setup_styles(self):
        """Configure custom ttk styles for professional appearance with card-based layout"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Color palette
        bg_color = '#f0f0f0' 
        frame_bg = '#ffffff' 
        accent_color = '#0078d4' 
        
        # Title label
        style.configure('Title.TLabel', font=('Segoe UI', 16, 'bold'), background=bg_color, foreground='#2c3e50')
        # Header label
        style.configure('Header.TLabel', font=('Segoe UI', 10, 'bold'), background=frame_bg)
        # Info label
        style.configure('Info.TLabel', font=('Segoe UI', 9), background=frame_bg, foreground='#555555')
        # Card frame
        style.configure('Card.TFrame', background=frame_bg, relief='flat', borderwidth=1)
        # Card label frame
        style.configure('Card.TLabelframe', background=frame_bg, relief='solid', borderwidth=1)
        # Card label frame label
        style.configure('Card.TLabelframe.Label', font=('Segoe UI', 10, 'bold'), foreground='#2c3e50', background=frame_bg)
    
    def setup_ui(self):
        """Setup the main UI layout with all cards and controls"""
        outer_container = tk.Frame(self.root, bg='#f0f0f0')
        outer_container.pack(fill="both", expand=True, padx=20, pady=20)
        
        main_container = tk.Frame(outer_container, bg='#f0f0f0')
        main_container.pack(fill="both", expand=True, anchor="center")
        
        # Title
        title_label = ttk.Label(main_container, text="🔄 SCADA Signal Synchronizer", style='Title.TLabel')
        title_label.pack(pady=(0, 20))
        
        # File Selection Card
        file_frame = ttk.LabelFrame(main_container, text="📁 Excel File Selection", style='Card.TLabelframe', padding=15)
        file_frame.pack(fill="x", pady=(0, 10))
        
        file_row = tk.Frame(file_frame, bg='#ffffff')
        file_row.pack(fill="x", pady=5)
        
        self.file_btn = tk.Button(file_row, text="Select Excel File", command=self.select_file,
                                  bg='#0078d4', fg='white', font=('Segoe UI', 9, 'bold'),
                                  relief='flat', padx=15, pady=8, cursor='hand2', width=15)
        self.file_btn.pack(side="left", padx=(0, 10))
        
        self.file_label = ttk.Label(file_row, text="No file selected", style='Info.TLabel')
        self.file_label.pack(side="left", fill="x")
        
        # Settings Card
        settings_frame = ttk.LabelFrame(main_container, text="⚙️ Settings", style='Card.TLabelframe', padding=15)
        settings_frame.pack(fill="x", pady=(0, 10))
        
        # Settings row 1
        settings_row1 = tk.Frame(settings_frame, bg='#ffffff')
        settings_row1.pack(fill="x", pady=5)
        
        self.update_tags_var = tk.BooleanVar(value=True)
        self.update_tags_btn = tk.Button(settings_row1, text="✓ Update tag names: True", 
                                        command=self.toggle_update_tags, width=25, bg='#0d6efd', fg='white',
                                        font=('Segoe UI', 9, 'bold'), relief='flat', padx=5, cursor='hand2')
        self.update_tags_btn.pack(side="left", padx=5)
        
        # Smart descriptions always enabled
        self.smart_desc_var = tk.BooleanVar(value=True)
        info_label = ttk.Label(settings_frame, text="✨ Smart descriptions enabled (Status suffixes not added)", style='Info.TLabel')
        info_label.pack(anchor="w", pady=8)
        
        # Log Card
        log_frame = ttk.LabelFrame(main_container, text="📝 Processing Log", style='Card.TLabelframe', padding=10)
        log_frame.pack(fill="both", expand=True, pady=(0, 10))
        
        log_container = tk.Frame(log_frame, bg='#ffffff')
        log_container.pack(fill="both", expand=True)
        
        self.log_text = tk.Text(log_container, height=20, wrap="word", bg='#f8f9fa', 
                               fg='#2c3e50', font=('Consolas', 9), relief='flat', padx=10, pady=10)
        self.log_text.pack(side="left", fill="both", expand=True)
        
        scrollbar = ttk.Scrollbar(log_container, command=self.log_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        # Button Frame
        button_frame = tk.Frame(main_container, bg='#f0f0f0', height=70)
        button_frame.pack(fill="x", pady=(0, 0))
        button_frame.pack_propagate(False)
        
        self.process_btn = tk.Button(
            button_frame, 
            text="▶️ Synchronize SCADA_SIGNAL", 
            command=self.show_save_dialog,
            bg='#28a745', 
            fg='white', 
            font=('Segoe UI', 11, 'bold'),
            relief='flat', 
            padx=40, 
            pady=10, 
            cursor='hand2',
            activebackground='#218838',
            activeforeground='white',
            width=20
        )
        self.process_btn.pack(expand=True)
    
    def log(self, message):
        """Add timestamped message to processing log and scroll to bottom"""
        self.log_text.insert("end", message + "\n")
        self.log_text.see("end")  # Auto-scroll to show latest message
        self.root.update()  # Force UI update to show message immediately
    
    def select_file(self):
        """Open file browser to select source Excel file"""
        file = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file:
            self.excel_file = file  # Store file path
            self.file_label.config(text=f"{Path(file).name}")  # Display filename in UI
            self.log(f"Selected file: {Path(file).name}")  # Log selection
    
    def toggle_update_tags(self):
        """Toggle 'Update tag names' setting and update button visual state"""
        self.update_tags_var.set(not self.update_tags_var.get())
        state = "True" if self.update_tags_var.get() else "False"
        emoji = "✓" if self.update_tags_var.get() else "✗"  # Emoji changes with state
        # Update button appearance: text, emoji, and color (blue for True, gray for False)
        self.update_tags_btn.config(text=f"{emoji} Update tag names: {state}",
                                   bg='#0d6efd' if self.update_tags_var.get() else '#6c757d')
    

    

    
    def format_signal_label(self, signal_type):
        """Format variant names for display in descriptions
        
        Converts CamelCase variants like 'HiAlarm' to 'HIGH ALARM' for descriptions.
        Handles special cases like 'HiHi' and 'LoLo' appropriately.
        Examples:
        - 'Status' → 'STATUS'
        - 'HiAlarm' → 'HIGH ALARM'
        - 'HiHi' → 'HIHI'
        - 'MyVariable' → 'MY VARIABLE'
        """
        if not signal_type:
            return ''
        s = str(signal_type)
        # Already all caps with underscores: just uppercase
        if s.replace('_', '').isupper():
            return s.replace('_', '').upper()
        
        # Convert underscores to spaces
        s = s.replace('_', ' ')
        # Insert spaces before uppercase letters (CamelCase splitting)
        s = re.sub(r'(?<!^)(?=[A-Z])', ' ', s)
        # Collapse multiple spaces
        s = ' '.join(s.split())
        # Convert to uppercase
        s = s.upper()
        # Fix HI HI → HIHI and LO LO → LOLO
        s = re.sub(r'\bHI\s+HI\b', 'HIHI', s)
        s = re.sub(r'\bLO\s+LO\b', 'LOLO', s)
        return s
    
    def normalize_for_matching(self, word):
        """Normalize word for case-insensitive and slash-to-dot matching"""
        return word.replace("/", ".").upper()  # Convert A/C to A.C and uppercase
    
    def show_save_dialog(self):
        """Show native save dialog for selecting output path and filename"""
        if not self.excel_file:
            messagebox.showerror("Error", "Please select an Excel file")
            return
        
        default_filename = f"{Path(self.excel_file).stem}_synchronized.xlsx"
        output_file = filedialog.asksaveasfilename(
            title="Save Synchronized File As",
            initialdir=str(Path(self.excel_file).parent),
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not output_file:
            self.log("✗ Save cancelled")
            return
        
        self.output_path = str(Path(output_file).parent)
        self.output_name = Path(output_file).stem
        self.process_file()
    
    def process_file(self):
        """Main processing function: synchronizes descriptions in 3 steps
        
        Step 1: Extract descriptions from all area sheets (AC, BILGES, etc.)
        Step 2: Update SCADA_SIGNAL rows with smart-formatted descriptions
        Step 3: Add missing tags from area sheets with intelligent variant detection
        
        Includes error handling and automatic folder opening on completion.
        """
        # Validation
        if not self.excel_file:
            messagebox.showerror("Error", "Please select an Excel file")
            return
        
        try:
            # Begin processing
            self.log("\n" + "=" * 80)
            self.log("Starting SCADA_SIGNAL Synchronization...")
            self.process_btn.config(state="disabled")  # Disable button during processing
            
            # Load Excel workbook
            wb = openpyxl.load_workbook(self.excel_file)
            
            # ==================== STEP 1: Extract Descriptions ====================
            self.log("\n📖 Step 1: Reading descriptions from area sheets...")
            self.log("-" * 80)
            
            area_descriptions = {}  # Dict: (sheet_name, tag_name) → description
            area_udt_types = {}  # Dict: (sheet_name, tag_name) → UDT Type
            
            # Process each sheet except SCADA_SIGNAL
            for sheet_name in wb.sheetnames:
                if sheet_name == "SCADA_SIGNAL":
                    continue  # Skip SCADA_SIGNAL, only read area sheets
                
                ws = wb[sheet_name]
                headers = {}
                for cell in ws[1]:
                    if cell.value:
                        headers[cell.value] = cell.column
                
                if "Tag Name" not in headers or "Description" not in headers:
                    continue
                
                tag_col = headers["Tag Name"]
                desc_col = headers["Description"]
                udt_col = headers.get("UDT Type", None)  # Get UDT Type if available
                
                sheet_count = 0
                for row in ws.iter_rows(min_row=2, values_only=False):
                    tag_name = row[tag_col - 1].value
                    desc = row[desc_col - 1].value
                    if tag_name and desc:
                        area_descriptions[(sheet_name, tag_name)] = desc
                        if udt_col:
                            udt_type = row[udt_col - 1].value
                            area_udt_types[(sheet_name, tag_name)] = udt_type
                        sheet_count += 1
                
                self.log(f"  {sheet_name}: {sheet_count} tags")
            
            self.log(f"✓ Total tags loaded: {len(area_descriptions)}")
            
            # ==================== STEP 2: Synchronize SCADA_SIGNAL ====================
            self.log("\n" + "=" * 80)
            self.log("🔄 Step 2: Synchronizing SCADA_SIGNAL sheet...")
            self.log("-" * 80)
            
            if "SCADA_SIGNAL" not in wb.sheetnames:
                self.log("ERROR: SCADA_SIGNAL sheet not found!")
                return
            
            ws = wb["SCADA_SIGNAL"]
            headers = {}
            for cell in ws[1]:
                if cell.value:
                    headers[cell.value] = cell.column
            
            if "Scada Tag Path" not in headers or "Description" not in headers:
                self.log("ERROR: Missing required columns in SCADA_SIGNAL")
                return
            
            tag_path_col = headers["Scada Tag Path"]
            db_col = headers.get("DB")
            desc_col = headers["Description"]
            
            updates = 0
            processed = 0
            
            self.log("\n  Processing SCADA_SIGNAL rows...")
            
            for row in ws.iter_rows(min_row=2, values_only=False):
                tag_path = row[tag_path_col - 1].value
                db = row[db_col - 1].value if db_col else None
                
                if not tag_path or not db:
                    continue
                
                path_parts = str(tag_path).split(".")
                if len(path_parts) < 3:
                    continue
                
                base_tag = path_parts[1]
                variant = path_parts[2]
                
                processed += 1
                
                key = (db, base_tag)
                if key not in area_descriptions:
                    continue
                
                area_desc = area_descriptions[key]
                if not isinstance(area_desc, str):
                    continue
                
                # Build description with smart formatting - never add .Status suffix
                formatted_variant = self.format_signal_label(variant)
                is_data_type = bool(re.fullmatch(r'[A-Z0-9]+', str(variant).replace('_', '')))
                
                # Only add suffix if it's not Status and not a data type
                if formatted_variant and not is_data_type and variant.upper() != 'STATUS':
                    expected_desc = f"{area_desc} {formatted_variant}".strip()
                else:
                    expected_desc = area_desc
                
                current_desc = row[desc_col - 1].value
                if not current_desc or current_desc.strip() != expected_desc.strip():
                    row[desc_col - 1].value = expected_desc
                    updates += 1
                    if updates <= 10:  # Show first 10 updates
                        self.log(f"  {db}.{base_tag}.{variant}")
            
            self.log(f"\n✨ Updated: {updates} rows")
            
            # ==================== Save Output ====================
            self.log("\n" + "=" * 80)
            self.log("💾 Saving file...")
            
            # Build output path and save with user-specified name
            output_file = str(Path(self.output_path) / f"{self.output_name}.xlsx")
            wb.save(output_file)  # Save to user-selected location
            wb.close()  # Close workbook
            
            self.log(f"\n✅ SUCCESS! File synchronized: {Path(output_file).name}")
            messagebox.showinfo("✅ Success", 
                              f"Synchronization complete! ✨\n\n"
                              f"📊 Processed: {processed} rows\n"
                              f"✏️ Updated: {updates} rows\n"
                              f"📁 Saved: {Path(output_file).name}\n"
                              f"📂 Location: {self.output_path}")
            
            # Open output folder
            output_folder = str(Path(output_file).parent)
            if sys.platform == 'win32':
                os.startfile(output_folder)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', output_folder])
            else:
                subprocess.Popen(['xdg-open', output_folder])
        
        except Exception as e:
            # Handle errors and display to user
            self.log(f"\n❌ ERROR: {str(e)}")
            messagebox.showerror("❌ Error", f"An error occurred:\n{str(e)}")
        
        finally:
            # Re-enable button after processing (success or failure)
            self.process_btn.config(state="normal")

# ====================================================================================
# Application Entry Point
# ====================================================================================
if __name__ == "__main__":
    root = tk.Tk()  # Create root window
    app = SCADASyncUI(root)  # Initialize application
    root.mainloop()  # Start event loop
